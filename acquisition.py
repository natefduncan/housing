import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from statistics import mean
from math import floor

#path = Path(os.path.dirname(os.path.abspath(__file__)))
path = Path("C:/Users/ndunc/Desktop/housing")
os.chdir(path)

from re_investment import mortgage, revenue, expenses, capex, returns, mc

os.chdir(path / "Acquisition")
#List Excel Files.
print("Excel Files:")
print("-" * 5)
files = os.listdir()
for i in range(0, len(files)):
    print(("%s: %s") % (str(i), files[i]))
print("-" * 5)

#Get file number. 
file_num = input("File Number: ")

#Get filename.
file_name = files[int(file_num)]

#Read in excel file.
wb = load_workbook(filename=file_name, data_only=True)

#Get Sheet
ass = wb["Assumptions"]
cap = wb["CAPEX"]
summary = wb["Summary"]

#Put in dataframes.
ass_df = pd.DataFrame(ass.values)
cap_df = pd.DataFrame(cap.values)
summary_df = pd.DataFrame(summary.values)

#Get Parameters. 
var = [i for j in ass_df[[0]].values.tolist() for i in j]

#Function to parse the assumption rows. 
def parse_assumptions(x):
    if x[1] != None:
        return float(x[1])
    else:
        if x[4] != None:
            return {"distribution" : x[2].lower(),
                    "value" : [float(x[3]), float(x[4])]}
        elif x[3] != None: 
            return {"distribution" : x[2].lower(),
                    "value" : [float(x[3])]}
        else:
            return None

#Mortgage Assumptions
loan_amnt = parse_assumptions(ass_df.iloc[var.index("Loan Amnt")])
interest = parse_assumptions(ass_df.iloc[var.index("Interest Rate")])
term = parse_assumptions(ass_df.iloc[var.index("Years")])
down_pmt = parse_assumptions(ass_df.iloc[var.index("Down Pmt")])


mortgage_parms = {"loan_amnt" : loan_amnt,
                  "interest" : interest,
                  "term" : term,
                  "down_pmt" : down_pmt,
                  }

#Occupancy Assumptions:
occupancy_length = parse_assumptions(ass_df.iloc[var.index("Occupancy Length")])
vacancy_length = parse_assumptions(ass_df.iloc[var.index("Vacancy Length")])

occupancy = {'type' : "simulation", #distribution = ["poisson"]
             'distribution' : "poisson", 
             'values' : [occupancy_length, vacancy_length] #Poisson=[occupancy_length, vacancy_length]
             }

#Revenue Assumptions
rent_mult = parse_assumptions(ass_df.iloc[var.index("Rent Multiple")])
rent_amnt = parse_assumptions(ass_df.iloc[var.index("Rent Amnt")])
increase = parse_assumptions(ass_df.iloc[var.index("Rent Increase")])
start = parse_assumptions(ass_df.iloc[var.index("Rent Start")])
change_rent_mnth = parse_assumptions(ass_df.iloc[var.index("Rent Change Mnth")])
change_rent_value = parse_assumptions(ass_df.iloc[var.index("Rent Change Value")]) 

revenue_parms = {"rent_mult" : rent_mult,
                "rent_amnt" : rent_amnt, 
                 "increase" : increase,
                 "start" : start,
                 "occupancy" : occupancy,
                 "change_rent" : [change_rent_mnth, change_rent_value]
                  }

#Expense Assumptions
property_tax = parse_assumptions(ass_df.iloc[var.index("Property Tax")])
property_management = parse_assumptions(ass_df.iloc[var.index("Property Management")])
utilities = parse_assumptions(ass_df.iloc[var.index("Utilities")])
opex_growth = parse_assumptions(ass_df.iloc[var.index("OPEX Growth")])

expense_parms = {"property_tax" : property_tax,
                  "property_management" : property_management,
                  "utilities" : utilities,
                  "opex_growth" : opex_growth
                  }

#CAPEX Assumptions
capex_months = [i for j in cap_df.iloc[1:][[1]].values for i in j]
capex_costs = [i for j in cap_df.iloc[1:][[2]].values for i in j]

capex_parms = {"months" : capex_months,
               "costs" : capex_costs}

#Get Number of Simulations
sims = input("Number of Simulations: ")

x = mc(int(sims), mortgage_parms, revenue_parms, expense_parms, capex_parms)

cfs = [i.cash_flow for i in x.sim]
df = pd.DataFrame(cfs)

book = load_workbook(file_name)
writer = pd.ExcelWriter(file_name, engine = 'openpyxl')
writer.book = book

df.to_excel(writer, sheet_name='Cash Flows')
    
writer.save()
writer.close()

